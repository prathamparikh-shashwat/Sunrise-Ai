import os
from typing import Dict, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from sheets import normalize_business_type, get_business_sheet_config

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
EXCEL_FILENAME = "consultant_data.xlsx"
EXCEL_PATH = os.path.join(DATA_DIR, EXCEL_FILENAME)

def sync_answers_to_local_sheet(business_type: str, answers: Dict[str, Any]) -> Dict[str, Any]:
    """
    Route questionnaire answers to a local Excel Spreadsheet.
    Stores all submissions in a single Excel file with separate sheets for each business type.
    """
    # 1. Normalize business type and get column headers/values
    normalized_type = normalize_business_type(business_type)
    headers, row_data = get_business_sheet_config(normalized_type, answers)
    
    # 2. Ensure data directory exists
    os.makedirs(DATA_DIR, exist_ok=True)
    
    # 3. Load or create workbook
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        # Remove the default sheet created by openpyxl to start fresh
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)
            
    # 4. Ensure sheet for business type exists
    sheet_title = normalized_type.capitalize()  # Tailoring, Retail, General
    if sheet_title not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_title)
        # Write headers
        ws.append(headers)
        # Style headers (bold)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
    else:
        ws = wb[sheet_title]
        # If sheet is empty for some reason, append headers
        if ws.max_row == 0 or (ws.max_row == 1 and not any(cell.value for cell in ws[1])):
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                
    # 5. Append row data
    ws.append(row_data)
    
    # 6. Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # 7. Save workbook
    wb.save(EXCEL_PATH)
    
    # Return path relative to project workspace
    relative_path = os.path.relpath(EXCEL_PATH, os.getcwd())
    
    return {
        "status": "success",
        "message": f"Successfully saved {normalized_type} questionnaire data locally.",
        "business_type": normalized_type,
        "sheet_name": sheet_title,
        "file_path": EXCEL_PATH,
        "relative_path": relative_path
    }
